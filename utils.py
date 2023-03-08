import pandas as pd
import openpyxl
import subprocess
from config import DATA_FILE
import sys
import logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')


def save_data(data, file_name):
    # Convertir les données en pandas DataFrame
    df = pd.DataFrame(data)
    
    # Écrire le DataFrame dans un fichier Excel
    writer = pd.ExcelWriter(file_name)
    df.to_excel(writer, index=False)
    writer.save()
    
    print(f"Les données ont été enregistrées dans le fichier {file_name}")
def add_entry_to_excel(entry):
    wb = openpyxl.load_workbook(DATA_FILE)
    sheet = wb.active
    sheet.append([entry['name'], entry['description'], entry['price']])
    wb.save(DATA_FILE)

def delete_entry_from_excel(entry):
    wb = openpyxl.load_workbook(DATA_FILE)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        if row[0] == entry['name'] and row[1] == entry['description'] and row[2] == entry['price']:
            sheet.delete_rows(row[0].row)
            break
    wb.save(DATA_FILE)

def get_entry_by_id(id):
    wb = openpyxl.load_workbook(DATA_FILE)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == id:
            return {'name': row[1], 'description': row[2], 'price': row[3]}
    return None

import pandas as pd


def add_entry_to_excel(data, file_path):
    df = pd.read_excel(file_path)
    df = df.append(data, ignore_index=True)
    df.to_excel(file_path, index=False)


def delete_entry_from_excel(entry_id, file_path):
    df = pd.read_excel(file_path)
    df = df[df['id'] != entry_id]
    df.to_excel(file_path, index=False)


def get_entry_by_id(entry_id, file_path):
    df = pd.read_excel(file_path)
    entry = df.loc[df['id'] == entry_id].to_dict(orient='records')
    return entry[0] if entry else None

def check_package(package_name):
    try:
        # Vérification si le package est déjà installé
        subprocess.check_call([sys.executable, "-c", f"import {package_name}"])
        print(f"{package_name} est déjà installé.")
    except subprocess.CalledProcessError:
        # Installation du package s'il n'est pas installé
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
        print(f"{package_name} a été installé.")
    else:
        # Mise à jour du package s'il est déjà installé
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", package_name])
        print(f"{package_name} a été mis à jour.")